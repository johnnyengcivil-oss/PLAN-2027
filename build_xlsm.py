"""Gera a pasta de trabalho do Sistema de Composições.

Constrói `Sistema_Composicoes.xlsx` com todas as abas, cabeçalhos, filtros
e textos de ajuda. Os módulos VBA ficam versionados em `vba/*.bas` e são
importados na pasta uma única vez (ver docs/INSTALACAO.md).

Motivo de a interface ser gerada por código: o .xlsm passa a ser um
artefato reconstruível a partir do repositório, e não um binário opaco.
Reconstruir é `python build_xlsm.py`.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent

# LAYOUT PADRÃO DAS ABAS — o VBA depende destas posições.
# Alterar aqui exige alterar as referências correspondentes em vba/*.bas.
LINHA_BANNER = 1        # 1-2: faixa de título (mescladas)
LINHA_CAMPOS = 3        # 3-5: filtros e campos de entrada (B3, B4, B5)
LINHA_MENSAGEM = 6      # 6:   linha de situação/contagem
LINHA_CABECALHO = 7     # 7:   cabeçalho da tabela
LINHA_DADOS = 8         # 8+:  dados

AZUL = "1F3864"
AZUL_CLARO = "D9E2F3"
CINZA = "404040"
CINZA_CLARO = "F2F2F2"
VERDE = "C6EFCE"
AMARELO = "FFEB9C"
LARANJA = "FFD9A0"
BRANCO = "FFFFFF"

TITULO = Font(name="Segoe UI", size=18, bold=True, color=BRANCO)
SUBTITULO = Font(name="Segoe UI", size=10, color=BRANCO)
SECAO = Font(name="Segoe UI", size=12, bold=True, color=AZUL)
ROTULO = Font(name="Segoe UI", size=10, bold=True)
NORMAL = Font(name="Segoe UI", size=10)
AJUDA = Font(name="Segoe UI", size=9, italic=True, color="595959")
NUMERO = Font(name="Segoe UI", size=14, bold=True, color=AZUL)

FUNDO_TITULO = PatternFill("solid", fgColor=AZUL)
FUNDO_CAMPO = PatternFill("solid", fgColor=BRANCO)
FUNDO_SECAO = PatternFill("solid", fgColor=AZUL_CLARO)
FUNDO_CINZA = PatternFill("solid", fgColor=CINZA_CLARO)

BORDA_CAMPO = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def faixa_titulo(ws, titulo: str, subtitulo: str, largura: int = 16) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=largura)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=largura)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font, c.fill = TITULO, FUNDO_TITULO
    c.alignment = Alignment(vertical="center", indent=1)
    d = ws.cell(row=2, column=1, value=subtitulo)
    d.font, d.fill = SUBTITULO, FUNDO_TITULO
    d.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, largura + 1):
        ws.cell(row=1, column=col).fill = FUNDO_TITULO
        ws.cell(row=2, column=col).fill = FUNDO_TITULO
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18


def campo(ws, linha: int, rotulo: str, valor: str = "",
          ajuda: str = "", coluna: int = 1) -> None:
    c = ws.cell(row=linha, column=coluna, value=rotulo)
    c.font = ROTULO
    v = ws.cell(row=linha, column=coluna + 1, value=valor)
    v.font, v.fill, v.border = NORMAL, FUNDO_CAMPO, BORDA_CAMPO
    if ajuda:
        a = ws.cell(row=linha, column=coluna + 2, value=ajuda)
        a.font = AJUDA


def secao(ws, linha: int, texto: str, largura: int = 16) -> None:
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura)
    c = ws.cell(row=linha, column=1, value=texto)
    c.font, c.fill = SECAO, FUNDO_SECAO
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[linha].height = 20


def indicador(ws, linha: int, rotulo: str, coluna: int = 2) -> None:
    c = ws.cell(row=linha, column=coluna, value=rotulo)
    c.font = ROTULO
    v = ws.cell(row=linha, column=coluna + 1)
    v.font = NUMERO
    v.alignment = Alignment(horizontal="left")


def larguras(ws, mapa: dict[int, float]) -> None:
    for col, largura in mapa.items():
        ws.column_dimensions[get_column_letter(col)].width = largura


# ------------------------------------------------------------------ abas

def aba_inicio(wb: Workbook):
    ws = wb.create_sheet("INICIO")
    faixa_titulo(ws, "BANCO PRÓPRIO DE COMPOSIÇÕES",
                 "Construção e manutenção das composições de preços da empresa "
                 "a partir das referências EDIF/INFRA")
    larguras(ws, {1: 3, 2: 34, 3: 18, 4: 20, 5: 14, 6: 46})

    # Instrução principal, antes de qualquer indicador: quem abre o
    # arquivo pela primeira vez precisa saber o que fazer, não quantos
    # registros existem.
    c = ws.cell(row=4, column=2,
                value="Clique em COMEÇAR, à direita  →")
    c.font = Font(name="Segoe UI", size=13, bold=True, color=AZUL)
    c = ws.cell(row=5, column=2,
                value="O assistente conduz você em 4 passos: escolher o serviço, "
                      "achar a referência EDIF/INFRA,")
    c.font = AJUDA
    c = ws.cell(row=6, column=2,
                value="conferir os insumos e gravar. Nada é gravado sem você mandar.")
    c.font = AJUDA

    secao(ws, 7, "SITUAÇÃO DO MOTOR")
    ws.cell(row=8, column=2, value="Motor Python").font = ROTULO
    ws.cell(row=9, column=2, value="Backend semântico").font = ROTULO

    secao(ws, 11, "SERVIÇOS DA EMPRESA")
    indicador(ws, 12, "Serviços cadastrados")
    indicador(ws, 13, "Serviços vinculados")
    indicador(ws, 14, "Serviços pendentes")

    secao(ws, 15, "COMPOSIÇÕES PRÓPRIAS")
    indicador(ws, 16, "Composições próprias")
    indicador(ws, 17, "  Completas")
    indicador(ws, 18, "  Pendentes")
    indicador(ws, 19, "  A revisar")

    secao(ws, 20, "VÍNCULOS E PENDÊNCIAS")
    indicador(ws, 21, "Materiais vinculados")
    indicador(ws, 22, "Equipamentos vinculados")
    indicador(ws, 23, "Pendências abertas")

    secao(ws, 24, "BASES CARREGADAS")
    indicador(ws, 25, "Materiais na base interna")
    indicador(ws, 26, "Composições de referência")

    ws.cell(row=28, column=2,
            value="As bases originais são abertas SOMENTE PARA LEITURA. "
                  "O sistema nunca grava, converte ou renomeia esses arquivos."
            ).font = AJUDA
    larguras(ws, {2: 34})
    return ws


def aba_servicos(wb: Workbook):
    ws = wb.create_sheet("SERVICOS")
    faixa_titulo(ws, "SERVIÇOS DA EMPRESA",
                 "Selecione um serviço e use BUSCAR CORRESPONDÊNCIA")
    larguras(ws, {1: 12, 2: 26, 3: 7, 4: 62, 5: 13, 6: 11, 7: 21, 8: 22, 9: 13, 10: 13})

    campo(ws, 3, "Família", "", "vazio = todas")
    campo(ws, 4, "Situação", "", "PENDENTES | VINCULADOS | vazio = todos")
    campo(ws, 5, "Palavra-chave", "", "filtra pela descrição")
    ws.cell(row=6, column=1, value="Resultado").font = ROTULO

    dv = DataValidation(type="list", formula1='"PENDENTES,VINCULADOS"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["B4"])
    return ws


def aba_correspondencia(wb: Workbook):
    ws = wb.create_sheet("CORRESPONDENCIA")
    faixa_titulo(ws, "CORRESPONDÊNCIA SERVIÇO → EDIF/INFRA",
                 "O algoritmo sugere; a escolha final é sempre do engenheiro")
    larguras(ws, {1: 5, 2: 14, 3: 12, 4: 20, 5: 9, 6: 12, 7: 60, 8: 7,
                  9: 13, 10: 10, 11: 11, 12: 10, 13: 10, 14: 10, 15: 70})

    campo(ws, 3, "Código do serviço", "", "")
    campo(ws, 4, "Origens", "EDIF,INFRA", "EDIF | INFRA | EDIF,INFRA")
    ws.cell(row=5, column=1, value="Serviço").font = ROTULO
    ws.cell(row=6, column=1, value="Detalhes").font = ROTULO

    dv = DataValidation(type="list", formula1='"EDIF,INFRA,EDIF INFRA"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["B4"])

    ws.cell(row=3, column=4,
            value="Nenhuma correspondência é gravada automaticamente, "
                  "mesmo com score de 100%. O algoritmo sugere; você decide."
            ).font = AJUDA
    return ws


def aba_composicao(wb: Workbook):
    ws = wb.create_sheet("COMPOSICAO")
    faixa_titulo(ws, "COMPOSIÇÃO",
                 "Composição de referência expandida e composição própria proposta")
    larguras(ws, {1: 14, 2: 14, 3: 58, 4: 7, 5: 14, 6: 13, 7: 13, 8: 9,
                  9: 24, 10: 12, 11: 46, 12: 8, 13: 13, 14: 24, 15: 40, 16: 60})
    campo(ws, 3, "Código do serviço", "", "")
    ws.cell(row=4, column=1, value="Referência").font = ROTULO
    ws.cell(row=5, column=1, value="Detalhes").font = ROTULO
    return ws


def aba_banco(wb: Workbook):
    ws = wb.create_sheet("BANCO_COMPOSICOES")
    faixa_titulo(ws, "BANCO DE COMPOSIÇÕES PRÓPRIAS",
                 "Composições construídas e validadas pela empresa")
    larguras(ws, {1: 15, 2: 16, 3: 56, 4: 7, 5: 11, 6: 12, 7: 13, 8: 11,
                  9: 13, 10: 13, 11: 13, 12: 13, 13: 14, 14: 13, 15: 44, 16: 19})
    campo(ws, 3, "Filtrar status", "",
          "COMPLETA | PENDENTE | REVISAR | DESATUALIZADA | vazio = todos")
    ws.cell(row=6, column=1, value="Resultado").font = ROTULO
    dv = DataValidation(
        type="list", formula1='"COMPLETA,PENDENTE,REVISAR,DESATUALIZADA"',
        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["B3"])
    return ws


def aba_pendencias(wb: Workbook):
    ws = wb.create_sheet("PENDENCIAS")
    faixa_titulo(ws, "CENTRAL DE PENDÊNCIAS",
                 "Uma composição nunca é descartada por causa de um insumo em aberto")
    larguras(ws, {1: 7, 2: 8, 3: 32, 4: 12, 5: 14, 6: 16, 7: 11, 8: 46, 9: 78, 10: 19})
    campo(ws, 3, "Filtrar tipo", "", "vazio = todos os tipos")
    ws.cell(row=4, column=1, value="Resumo").font = ROTULO
    ws.cell(row=6, column=1, value="Resultado").font = ROTULO
    return ws


def aba_configuracao(wb: Workbook):
    ws = wb.create_sheet("CONFIGURACAO")
    faixa_titulo(ws, "CONFIGURAÇÃO",
                 "Identificação das bases, política de preço e pesos do algoritmo")
    larguras(ws, {1: 26, 2: 58, 3: 40, 4: 50})

    campo(ws, 4, "Pasta do sistema", "", "derivada de ThisWorkbook.Path")
    campo(ws, 5, "Pasta das bases", "", "")
    campo(ws, 6, "Backend semântico", "", "")
    campo(ws, 7, "Política de preço", "VALOR_APROVADO",
          "VALOR_APROVADO | ULTIMO | MAX | MEDIA_RECENTE | MEDIANA")
    dv = DataValidation(
        type="list",
        formula1='"VALOR_APROVADO,ULTIMO,MAX,MEDIA_RECENTE,MEDIANA"',
        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["B7"])

    secao(ws, 9, "BASES IDENTIFICADAS", largura=4)
    ws.cell(row=10, column=1, value="PAPEL").font = ROTULO
    ws.cell(row=10, column=2, value="ARQUIVO DETECTADO").font = ROTULO
    ws.cell(row=10, column=3, value="FORÇAR ARQUIVO (opcional)").font = ROTULO

    ws.cell(row=17, column=1,
            value="A identificação de EDIF, INFRA e AUX é feita pelo CONTEÚDO do "
                  "arquivo — o título institucional e o nome da aba — e não pelo "
                  "nome, ordem ou tamanho. Preencha a coluna C apenas se quiser "
                  "sobrepor a detecção; a escolha fica gravada em config.json."
            ).font = AJUDA
    ws.cell(row=19, column=1,
            value="Para sobrepor, escreva em C13 / C14 / C15 o NOME DO ARQUIVO "
                  "a tratar como EDIF / INFRA / AUX e use SALVAR CONFIGURAÇÃO."
            ).font = AJUDA
    return ws


def aba_log(wb: Workbook):
    ws = wb.create_sheet("LOG")
    faixa_titulo(ws, "LOG DE AUDITORIA",
                 "Data, usuário, ação, referência, score e alterações")
    larguras(ws, {1: 21, 2: 16, 3: 24, 4: 22, 5: 20, 6: 72, 7: 9, 8: 6})
    ws.cell(row=3, column=1, value="Registros").font = ROTULO
    return ws


def aba_ajuda(wb: Workbook):
    ws = wb.create_sheet("AJUDA")
    faixa_titulo(ws, "COMO USAR", "Fluxo de trabalho e o que o sistema garante")
    larguras(ws, {1: 4, 2: 116})
    passos = [
        ("O CAMINHO NORMAL — USE O ASSISTENTE", ""),
        ("", "Na aba INÍCIO, clique em COMEÇAR — ASSISTENTE DE COMPOSIÇÃO. "
             "Ele conduz tudo em quatro passos, explicando cada um. As abas "
             "abaixo continuam disponíveis, mas você não precisa delas para "
             "o trabalho do dia a dia."),
        ("1", "ESCOLHER O SERVIÇO — digite parte do nome para filtrar os 949 "
              "serviços. Duplo clique já avança."),
        ("2", "ESCOLHER A REFERÊNCIA — o sistema mostra os candidatos EDIF e "
              "INFRA ordenados por semelhança, e o painel da direita explica "
              "cada score em barras: texto, sentido, termos-chave, unidade e "
              "técnico, com o que pesou a favor e contra."),
        ("3", "CONFERIR OS INSUMOS — clique num insumo e edite abaixo: trocar "
              "o item da empresa, mudar o coeficiente (botões −10% e +10%, ou "
              "digite direto) ou deixá-lo fora do custo. O total é recalculado "
              "a cada mudança."),
        ("4", "GRAVAR — a composição entra no banco próprio e os itens que você "
              "escolheu viram vínculos validados, sugeridos automaticamente nas "
              "próximas composições."),
        ("", ""),
        ("AS ABAS, PARA CONSULTA E CASOS ESPECIAIS", ""),
        ("SERVIÇOS", "lista completa, com filtros e análise em lote"),
        ("CORRESPONDÊNCIA", "a busca do passo 2, em formato de planilha"),
        ("COMPOSIÇÃO", "a composição de referência expandida, hierárquica e "
                       "consolidada, com o caminho de cada auxiliar"),
        ("BANCO_COMPOSIÇÕES", "tudo o que já foi gravado, com custos e situação"),
        ("PENDÊNCIAS", "o que ficou em aberto, das mais críticas para as menos"),
        ("CONFIGURAÇÃO", "identificação das bases e política de preço"),
        ("LOG", "quem fez o quê, quando, com qual score"),
        ("", ""),
        ("O QUE O SISTEMA GARANTE", ""),
        ("•", "As bases originais nunca são alteradas: leitura pura, "
              "transformação em memória e no banco SQLite local."),
        ("•", "Nenhum vínculo é confirmado automaticamente, nem com score de "
              "100%. O algoritmo sugere; o engenheiro decide."),
        ("•", "A mão de obra da composição de referência NÃO é somada ao custo "
              "próprio — o serviço interno já representa a execução. Ela fica "
              "registrada para rastreabilidade."),
        ("•", "Conversões de unidade são determinísticas. Quando dependem do "
              "produto e não podem ser deduzidas, viram pendência em vez de "
              "chute — e você pode digitar o coeficiente na mão."),
        ("•", "Cada item guarda código, descrição, unidade e coeficiente dos "
              "dois lados, o fator de conversão, o score, a data e o usuário."),
        ("•", "Um insumo problemático não descarta a composição: os demais são "
              "montados e o item aberto vira pendência."),
        ("•", "Vínculos já validados são reaproveitados e sinalizados como "
              "VÍNCULO VALIDADO, distintos de SUGESTÃO AUTOMÁTICA."),
        ("•", "Funciona 100% local, sem API paga e sem exigir privilégio de "
              "administrador."),
    ]
    linha = 4
    for marcador, texto in passos:
        if texto == "" and marcador:
            secao(ws, linha, marcador, largura=2)
        else:
            ws.cell(row=linha, column=1, value=marcador).font = ROTULO
            c = ws.cell(row=linha, column=2, value=texto)
            c.font = NORMAL
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[linha].height = 28
        linha += 1
    return ws


def construir(destino: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    aba_inicio(wb)
    aba_servicos(wb)
    aba_correspondencia(wb)
    aba_composicao(wb)
    aba_banco(wb)
    aba_pendencias(wb)
    aba_configuracao(wb)
    aba_log(wb)
    aba_ajuda(wb)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"   # mantém a faixa de título visível
    wb.active = 0
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


def main(argv: list[str] | None = None) -> int:
    argv = argv or sys.argv[1:]
    destino = Path(argv[0]) if argv else RAIZ / "Sistema_Composicoes.xlsx"
    caminho = construir(destino)
    modulos = sorted((RAIZ / "vba").glob("*.bas"))
    print(f"Pasta gerada: {caminho}")
    print(f"Abas: {', '.join(ws.title for ws in __import__('openpyxl').load_workbook(caminho).worksheets)}")
    print(f"\nMódulos VBA a importar ({len(modulos)}):")
    for m in modulos:
        print(f"  vba/{m.name}")
    print("\nPróximo passo: docs/INSTALACAO.md (importar os módulos e salvar "
          "como Sistema_Composicoes.xlsm).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
