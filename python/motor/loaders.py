"""Carregadores das bases (itens 40, 41, 53, 54).

REGRA ABSOLUTA: os arquivos originais são abertos SOMENTE PARA LEITURA.
Nada é gravado, convertido, renomeado ou salvo por cima. Toda a
transformação acontece em memória e termina no SQLite.

Cabeçalhos não são presumidos na linha 1: `localizar_cabecalho` procura a
linha que contém os rótulos esperados (item 41).
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator, Sequence

from .normalize import (
    limpar_espacos,
    normalizar_codigo,
    normalizar_texto,
    normalizar_unidade,
    para_numero,
    sem_acento,
)

PAPEIS_REFERENCIA = ("EDIF", "INFRA", "AUX")


# ------------------------------------------------------------------ utilidades

_CACHE_INFO: dict[tuple[str, int, float], dict[str, Any]] = {}


def hash_arquivo(caminho: Path) -> str:
    """SHA-256 do arquivo, para detectar alteração entre execuções (item 54)."""
    h = hashlib.sha256()
    with open(caminho, "rb") as fh:          # 'rb' — leitura pura
        for bloco in iter(lambda: fh.read(1 << 20), b""):
            h.update(bloco)
    return h.hexdigest()


def info_arquivo(caminho: Path) -> dict[str, Any]:
    """Identidade do arquivo. Memoizada por (caminho, tamanho, mtime) —
    a mesma importação consulta o hash mais de uma vez."""
    st = caminho.stat()
    chave = (str(caminho.resolve()), st.st_size, st.st_mtime)
    memo = _CACHE_INFO.get(chave)
    if memo is not None:
        return dict(memo)
    info = {
        "nome_arquivo": caminho.name,
        "hash_sha256": hash_arquivo(caminho),
        "tamanho_bytes": st.st_size,
        "data_modificacao": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
    }
    _CACHE_INFO[chave] = dict(info)
    return info


def _rotulo(valor: object) -> str:
    return sem_acento(limpar_espacos(valor)).upper()


def localizar_cabecalho(
    linhas: Sequence[Sequence[Any]],
    obrigatorios: Sequence[str],
    limite: int = 30,
) -> int | None:
    """Índice da linha de cabeçalho que contém todos os rótulos obrigatórios.

    Não assume A1 (item 41). A comparação ignora acento, caixa e
    no-break spaces — o cabeçalho da base de serviços usa `\\xa0`.
    """
    alvos = [_rotulo(o) for o in obrigatorios]
    for i, linha in enumerate(linhas[:limite]):
        celulas = [_rotulo(c) for c in linha]
        if all(any(alvo in celula for celula in celulas if celula) for alvo in alvos):
            return i
    return None


def mapear_colunas(linha: Sequence[Any], nomes: dict[str, Sequence[str]]) -> dict[str, int]:
    """Mapeia campo lógico -> índice de coluna, pelos rótulos aceitos."""
    celulas = [_rotulo(c) for c in linha]
    saida: dict[str, int] = {}
    for campo, aceitos in nomes.items():
        for idx, celula in enumerate(celulas):
            if not celula:
                continue
            if any(_rotulo(a) == celula for a in aceitos):
                saida[campo] = idx
                break
        else:
            for idx, celula in enumerate(celulas):
                if celula and any(_rotulo(a) in celula for a in aceitos):
                    saida[campo] = idx
                    break
    return saida


def _ler_xls(caminho: Path) -> tuple[str, list[list[Any]]]:
    """Lê a primeira aba de um .xls legado (BIFF). Somente leitura."""
    import xlrd

    livro = xlrd.open_workbook(str(caminho), formatting_info=False, on_demand=False)
    try:
        aba = livro.sheet_by_index(0)
        linhas = [[aba.cell_value(r, c) for c in range(aba.ncols)]
                  for r in range(aba.nrows)]
        return aba.name, linhas
    finally:
        livro.release_resources()


def _ler_xlsx(caminho: Path, aba: str | None = None) -> tuple[str, list[list[Any]]]:
    """Lê uma aba de um .xlsx. read_only=True não grava nada no arquivo."""
    import openpyxl

    wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
    try:
        ws = wb[aba] if aba and aba in wb.sheetnames else wb[wb.sheetnames[0]]
        largura = 0
        brutas = []
        for linha in ws.iter_rows(values_only=True):
            valores = list(linha)
            largura = max(largura, len(valores))
            brutas.append(valores)
        linhas = [v + [None] * (largura - len(v)) for v in brutas]
        return ws.title, linhas
    finally:
        wb.close()


def ler_planilha(caminho: Path) -> tuple[str, list[list[Any]]]:
    """Despacha para o leitor adequado ao formato (item 40)."""
    sufixo = caminho.suffix.lower()
    if sufixo == ".xls":
        return _ler_xls(caminho)
    if sufixo in {".xlsx", ".xlsm"}:
        return _ler_xlsx(caminho)
    raise ValueError(f"Formato não suportado: {caminho.name}")


# ------------------------------------------------- detecção EDIF / INFRA / AUX

_MARCADORES = {
    "AUX": (r"COMPOSIC\w*\s+AUXILIAR", r"AUXILIARES"),
    "EDIF": (r"EDIFICAC\w*", r"\bEDIF\b"),
    "INFRA": (r"INFRAESTRUTURA\s+URBANA", r"\bINFRA\b"),
}


def detectar_origem(nome_aba: str, linhas: Sequence[Sequence[Any]],
                    nome_arquivo: str = "") -> tuple[str, str]:
    """Identifica se um .xls é EDIF, INFRA ou AUX (item 4).

    Prioriza o CONTEÚDO do arquivo — título institucional das primeiras
    linhas e nome da aba. O nome do arquivo é pista secundária. Nunca
    infere por ordem ou tamanho. Devolve (origem, como_foi_detectado).
    """
    cabecalho = " | ".join(
        _rotulo(c) for linha in linhas[:8] for c in linha if limpar_espacos(c)
    )
    aba = _rotulo(nome_aba)

    # AUX antes de EDIF/INFRA: o título das auxiliares não cita nenhum dos dois.
    for origem in ("AUX", "EDIF", "INFRA"):
        for padrao in _MARCADORES[origem]:
            if re.search(padrao, cabecalho):
                return origem, f"conteúdo do arquivo (linha de título: /{padrao}/)"
    for origem in ("AUX", "EDIF", "INFRA"):
        for padrao in _MARCADORES[origem]:
            if re.search(padrao, aba):
                return origem, f"nome da aba ({nome_aba!r})"
    arq = _rotulo(nome_arquivo)
    for origem in ("AUX", "EDIF", "INFRA"):
        for padrao in _MARCADORES[origem]:
            if re.search(padrao, arq):
                return origem, f"nome do arquivo ({nome_arquivo!r})"
    return "", "não identificado"


_RE_DATA_BASE = re.compile(r"DATA[\s\-]*BASE\s*:?\s*([A-Z]{3,10}\s*/\s*\d{2,4})")


def detectar_data_base(linhas: Sequence[Sequence[Any]]) -> str:
    """Extrai a data-base do cabeçalho (item 55): 'JAN/2026'."""
    for linha in linhas[:8]:
        for celula in linha:
            texto = _rotulo(celula)
            m = _RE_DATA_BASE.search(texto)
            if m:
                return re.sub(r"\s+", "", m.group(1))
    return ""


# ------------------------------------------------- classificação de insumos

def classificar_insumo(codins: str, descricao: str, unidade: str,
                       codigos_auxiliares: set[str]) -> tuple[str, str]:
    """Classifica um insumo da referência (item 19).

    Usa a estrutura da própria base — faixa do CODINS e resolução contra a
    tabela de auxiliares — e só recorre ao texto como desempate. Ver
    diagnóstico §3 para a validação estatística destas faixas.
    """
    cod = normalizar_codigo(codins)
    und = normalizar_unidade(unidade)

    if cod in codigos_auxiliares:
        return "COMPOSICAO_AUXILIAR", "AUX"

    # Mão de obra: CODINS de 4 dígitos (faixas 1xxx/2xxx), 95/95 com unidade H.
    if len(cod) == 4 and cod.isdigit():
        return "MAO_DE_OBRA", ""

    # Equipamento: faixa 94xxx. Duas exceções conhecidas resolvidas pela
    # unidade — equipamento é cobrado por hora; item unitário é material.
    if cod.startswith("94") and len(cod) == 5 and cod.isdigit():
        if und == "UN":
            return "MATERIAL", ""      # ex.: 94531 CÂMARA DE PNEU
        return "EQUIPAMENTO", ""

    # Reforço textual apenas quando a faixa não decidiu.
    if und == "H" and re.search(r"\(SGSP\)", normalizar_texto(descricao)):
        return "MAO_DE_OBRA", ""

    return "MATERIAL", ""


# ------------------------------------------------- escopo do serviço interno

_PADROES_ESCOPO = (
    ("MAO_DE_OBRA", r"\bMAO DE OBRA\b|\bM\.?\s?O\.?\s+ESPECIALIZ|\bAPENAS MAO\b"),
    ("FORNEC_E_INSTAL",
     r"\bFORNECIMENTO E (INSTALACAO|COLOCACAO|ASSENTAMENTO|MONTAGEM|APLICACAO)\b"
     r"|\bFORNEC\.? E INST"),
    ("FORNECIMENTO", r"^FORNECIMENTO\b|\bFORNECIMENTO DE\b"),
    ("DEMOLICAO_REMOCAO", r"\bDEMOLICAO\b|\bRETIRADA\b|\bREMOCAO\b|\bDESMONTAGEM\b"),
    ("LOCACAO", r"\bLOCACAO\b|\bALUGUEL\b"),
)


def classificar_escopo(descricao: str) -> str:
    """Classifica o escopo do serviço interno (ver diagnóstico §5.1).

    Determina se a composição própria pode receber os materiais da
    referência sem dupla contagem. A ordem dos padrões importa: "MÃO DE
    OBRA ESPECIALIZADA PARA INSTALAÇÃO" é mão de obra, não instalação.
    """
    t = normalizar_texto(descricao)
    for escopo, padrao in _PADROES_ESCOPO:
        if re.search(padrao, t):
            return escopo
    return "EXECUCAO_INDEFINIDO"


# ------------------------------------------------- histórico de preços

_RE_PRECO_HIST = re.compile(
    r"R\$\s*([\d.,]+)\s*DATA\s*:\s*(\d{2})/(\d{2})/(\d{2,4})", re.IGNORECASE)


def parsear_historico_precos(texto: object) -> list[tuple[str | None, float]]:
    """Extrai o histórico de `ULTIMO_PRECO` (ver diagnóstico §6.2).

    Devolve [(data ISO, preço)] na ordem em que aparece — que nas bases
    auditadas é sempre da cotação mais recente para a mais antiga.
    """
    if texto is None:
        return []
    saida: list[tuple[str | None, float]] = []
    for m in _RE_PRECO_HIST.finditer(str(texto)):
        preco = para_numero(m.group(1))
        if preco is None:
            continue
        dia, mes, ano = int(m.group(2)), int(m.group(3)), int(m.group(4))
        if ano < 100:
            ano += 2000
        try:
            data = datetime(ano, mes, dia).date().isoformat()
        except ValueError:
            data = None
        saida.append((data, preco))
    return saida


def preco_vigente(historico: list[tuple[str | None, float]],
                  valor_coluna: float | None,
                  politica: str) -> float | None:
    """Aplica a política de preço configurada (item 52).

    O padrão `VALOR_APROVADO` usa a coluna VALOR — que a auditoria provou
    ser o máximo do histórico em 10.610/10.610 registros.
    """
    precos = [p for _, p in historico]
    if politica == "ULTIMO":
        return precos[0] if precos else valor_coluna
    if politica == "MAX":
        return max(precos) if precos else valor_coluna
    if politica == "MEDIA_RECENTE":
        return sum(precos) / len(precos) if precos else valor_coluna
    if politica == "MEDIANA" and precos:
        ordenados = sorted(precos)
        meio = len(ordenados) // 2
        if len(ordenados) % 2:
            return ordenados[meio]
        return (ordenados[meio - 1] + ordenados[meio]) / 2
    return valor_coluna


# ------------------------------------------------------------------ resultados

@dataclass
class ResultadoCarga:
    papel: str
    origem: str = ""
    registros: int = 0
    insumos: int = 0
    data_base: str = ""
    detectado_por: str = ""
    aba: str = ""
    avisos: list[str] = field(default_factory=list)
    info: dict[str, Any] = field(default_factory=dict)


# ------------------------------------------------------------------ serviços

COLUNAS_SERVICOS = {
    "familia": ("FAMILIA",),
    "codigo": ("CODIGO",),
    "unidade": ("UN", "UNID", "UNIDADE"),
    "descricao": ("DESCRICAO",),
    "aprovado": ("PRECO APROVADO", "APROVADO"),
    "valor": ("VALOR", "PRECO"),
}


def carregar_servicos(caminho: Path) -> tuple[ResultadoCarga, list[dict[str, Any]]]:
    """Carrega a base de serviços/mão de obra da empresa.

    Descarta as linhas separadoras de família (nome da família sem código)
    e classifica o escopo de cada serviço.
    """
    aba, linhas = ler_planilha(caminho)
    idx = localizar_cabecalho(linhas, ("CODIGO", "DESCRICAO"))
    res = ResultadoCarga(papel="SERVICOS", aba=aba, info=info_arquivo(caminho))
    if idx is None:
        res.avisos.append("Cabeçalho não localizado (esperados CÓDIGO e DESCRIÇÃO).")
        return res, []

    col = mapear_colunas(linhas[idx], COLUNAS_SERVICOS)
    faltando = {"codigo", "descricao"} - set(col)
    if faltando:
        res.avisos.append(f"Colunas obrigatórias ausentes: {sorted(faltando)}")
        return res, []

    def cel(linha: Sequence[Any], campo: str) -> Any:
        i = col.get(campo)
        return linha[i] if i is not None and i < len(linha) else None

    registros: list[dict[str, Any]] = []
    separadoras = 0
    for n, linha in enumerate(linhas[idx + 1:], start=idx + 2):
        codigo = normalizar_codigo(cel(linha, "codigo"))
        if not codigo:
            if limpar_espacos(cel(linha, "familia")):
                separadoras += 1
            continue
        descricao = limpar_espacos(cel(linha, "descricao"))
        if not descricao:
            res.avisos.append(f"Linha {n}: código {codigo} sem descrição — ignorado.")
            continue
        unidade_orig = limpar_espacos(cel(linha, "unidade"))
        valor_bruto = cel(linha, "valor")
        aprovado = _rotulo(cel(linha, "aprovado")).startswith("SIM")
        registros.append({
            "codigo": codigo,
            "familia": limpar_espacos(cel(linha, "familia")),
            "unidade": normalizar_unidade(unidade_orig),
            "unidade_orig": unidade_orig,
            "descricao": descricao,
            "descricao_norm": normalizar_texto(descricao),
            "preco": para_numero(valor_bruto),
            "preco_texto": limpar_espacos(valor_bruto),
            "preco_aprovado": 1 if aprovado else 0,
            "escopo": classificar_escopo(descricao),
            "linha_origem": n,
        })
    res.registros = len(registros)
    if separadoras:
        res.avisos.append(f"{separadoras} linhas separadoras de família ignoradas.")
    return res, registros


# ------------------------------------------------------------------ materiais

COLUNAS_MATERIAIS = {
    "item": ("ITEM",),
    "codigo": ("CODIGO_MATERIAL", "CODIGO"),
    "familia": ("FAMILIA",),
    "unidade": ("UNIDADE", "UN"),
    "descricao": ("MATERIAL", "DESCRICAO"),
    "historico": ("ULTIMO_PRECO",),
    "valor": ("VALOR",),
}

# Famílias que representam equipamento/ferramenta, não material de aplicação
# (ver diagnóstico §6.1).
FAMILIAS_EQUIPAMENTO = {"EQUIPAMENTOS", "FERRAMENTAS", "LOCACAO", "VEICULOS"}


def carregar_materiais(caminho: Path, politica_preco: str = "VALOR_APROVADO"
                       ) -> tuple[ResultadoCarga, list[dict[str, Any]]]:
    """Carrega a base de materiais da empresa, com histórico de preços."""
    aba, linhas = ler_planilha(caminho)
    idx = localizar_cabecalho(linhas, ("CODIGO", "MATERIAL"))
    res = ResultadoCarga(papel="MATERIAIS", aba=aba, info=info_arquivo(caminho))
    if idx is None:
        res.avisos.append("Cabeçalho não localizado (esperados CODIGO_MATERIAL e MATERIAL).")
        return res, []

    col = mapear_colunas(linhas[idx], COLUNAS_MATERIAIS)
    faltando = {"codigo", "descricao"} - set(col)
    if faltando:
        res.avisos.append(f"Colunas obrigatórias ausentes: {sorted(faltando)}")
        return res, []

    def cel(linha: Sequence[Any], campo: str) -> Any:
        i = col.get(campo)
        return linha[i] if i is not None and i < len(linha) else None

    registros: list[dict[str, Any]] = []
    for n, linha in enumerate(linhas[idx + 1:], start=idx + 2):
        codigo = normalizar_codigo(cel(linha, "codigo"))
        if not codigo:
            continue
        descricao = limpar_espacos(cel(linha, "descricao"))
        if not descricao:
            continue
        familia = limpar_espacos(cel(linha, "familia"))
        unidade_orig = limpar_espacos(cel(linha, "unidade"))
        historico = parsear_historico_precos(cel(linha, "historico"))
        valor = para_numero(cel(linha, "valor"))
        tipo = ("EQUIPAMENTO"
                if sem_acento(familia).upper() in FAMILIAS_EQUIPAMENTO
                else "MATERIAL")
        registros.append({
            "codigo": codigo,
            "item": int(para_numero(cel(linha, "item")) or 0) or None,
            "familia": familia,
            "unidade": normalizar_unidade(unidade_orig),
            "unidade_orig": unidade_orig,
            "descricao": descricao,
            "descricao_norm": normalizar_texto(descricao),
            "tipo_item": tipo,
            "preco": preco_vigente(historico, valor, politica_preco),
            "preco_valor": valor,
            "preco_ultimo": historico[0][1] if historico else None,
            "data_ultimo": historico[0][0] if historico else None,
            "politica_preco": politica_preco,
            "fonte": "MATERIAIS",
            "linha_origem": n,
            "_historico": historico,
        })
    res.registros = len(registros)
    return res, registros


# ------------------------------------------------------- referência EDIF/INFRA/AUX

def carregar_referencia(
    caminho: Path,
    codigos_auxiliares: set[str] | None = None,
    origem_forcada: str = "",
) -> tuple[ResultadoCarga, list[dict[str, Any]], list[dict[str, Any]]]:
    """Carrega uma base de composições de referência.

    O layout de 10 colunas é idêntico em EDIF, INFRA e AUX — um único
    parser atende às três. A distinção entre linha de composição e linha
    de insumo é estrutural: composição tem CÓDIGO na coluna 0, insumo tem
    CODINS na coluna 2 e pertence à última composição vista.
    """
    aba, linhas = ler_planilha(caminho)
    res = ResultadoCarga(papel="REFERENCIA", aba=aba, info=info_arquivo(caminho))

    if origem_forcada:
        res.origem, res.detectado_por = origem_forcada, "configuração do usuário"
    else:
        res.origem, res.detectado_por = detectar_origem(aba, linhas, caminho.name)
    if not res.origem:
        res.avisos.append(
            "Origem (EDIF/INFRA/AUX) não identificada pelo conteúdo. "
            "Defina manualmente na aba CONFIGURAÇÃO.")
        return res, [], []

    res.data_base = detectar_data_base(linhas)
    idx = localizar_cabecalho(linhas, ("CODIGO", "NOME DO SERVICO"))
    if idx is None:
        idx = localizar_cabecalho(linhas, ("CODIGO",)) or 0
    # A linha seguinte ao cabeçalho traz os subtítulos (CODINS/NOMINS/COEF.).
    inicio = idx + 1
    if inicio < len(linhas) and any(
            "CODINS" in _rotulo(c) for c in linhas[inicio]):
        inicio += 1

    codigos_auxiliares = codigos_auxiliares or set()
    composicoes: list[dict[str, Any]] = []
    insumos: list[dict[str, Any]] = []
    atual: str | None = None
    seq = 0
    orfaos = 0

    for n, linha in enumerate(linhas[inicio:], start=inicio + 1):
        def c(i: int) -> Any:
            return linha[i] if i < len(linha) else None

        codigo = normalizar_codigo(c(0))
        codins = normalizar_codigo(c(2))

        if codigo and codigo.isdigit():
            descricao = limpar_espacos(c(1))
            unidade_orig = limpar_espacos(c(7))
            atual, seq = codigo, 0
            composicoes.append({
                "origem": res.origem,
                "codigo": codigo,
                "descricao": descricao,
                "descricao_norm": normalizar_texto(descricao),
                "unidade": normalizar_unidade(unidade_orig),
                "unidade_orig": unidade_orig,
                "custo_total": para_numero(c(9)),
                "data_base": res.data_base,
                "linha_origem": n,
            })
        elif codins and codins.isdigit():
            if atual is None:
                orfaos += 1
                continue
            descricao = limpar_espacos(c(3))
            unidade_orig = limpar_espacos(c(4))
            classe, origem_aux = classificar_insumo(
                codins, descricao, unidade_orig, codigos_auxiliares)
            seq += 1
            insumos.append({
                "origem": res.origem,
                "codigo": atual,
                "seq": seq,
                "codins": codins,
                "descricao": descricao,
                "descricao_norm": normalizar_texto(descricao),
                "unidade": normalizar_unidade(unidade_orig),
                "unidade_orig": unidade_orig,
                "custo_unitario": para_numero(c(5)),
                "coeficiente": para_numero(c(6)) or 0.0,
                "valor_parcial": para_numero(c(8)),
                "classe": classe,
                "origem_aux": origem_aux,
                "linha_origem": n,
            })

    res.registros = len(composicoes)
    res.insumos = len(insumos)
    if orfaos:
        res.avisos.append(f"{orfaos} linhas de insumo sem composição-pai ignoradas.")
    return res, composicoes, insumos


def ler_codigos_auxiliares(caminho: Path) -> set[str]:
    """Códigos da base de auxiliares — necessários ANTES de classificar
    os insumos de EDIF/INFRA, já que uma auxiliar é reconhecida por
    aparecer como CÓDIGO nessa base."""
    aba, linhas = ler_planilha(caminho)
    codigos: set[str] = set()
    for linha in linhas:
        codigo = normalizar_codigo(linha[0] if linha else None)
        if codigo and codigo.isdigit():
            codigos.add(codigo)
    return codigos
